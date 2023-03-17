import matplotlib.pyplot
import openpyxl as xl
import matplotlib.pyplot as plt
import numpy as np


# returns a ordered list from a colum ignoring the first element (colum name)
def get_index(column, ws):
    col = []
    for i in range(2, len(ws[column])):
        col.append(float(ws.cell(row=i, column=(ord(column) - 64)).value))
    col.sort()
    return col


# the object that will containg all the data in a season
class Season:
    def __init__(self, workbook):
        self.workbook = workbook
        print("loading workbook: " + workbook)
        wb = xl.load_workbook(filename='workbooks\\' + workbook + ".xlsx")  # load workbook from file
        ws = wb.active

        # define all the indexes into ordered vectors
        self.FG = get_index('E', ws)
        self.FG_attempted = get_index('F', ws)
        self.FG_rate = get_index('G', ws)
        self.ThreeP = get_index('H', ws)
        self.ThreeP_attempted = get_index('I', ws)
        self.ThreeP_rate = get_index('J', ws)
        self.TwoP = get_index('K', ws)
        self.TwoP_attempted = get_index('L', ws)
        self.TwoP_rate = get_index('M', ws)
        self.FT = get_index('N', ws)
        self.FT_attempted = get_index('O', ws)
        self.FT_rate = get_index('P', ws)
        self.OR = get_index('Q', ws)
        self.DR = get_index('R', ws)
        self.AST = get_index('T', ws)
        self.STL = get_index('U', ws)
        self.BLK = get_index('V', ws)
        self.TOV = get_index('W', ws)
        self.PF = get_index('X', ws)
        self.PTS = get_index('Y', ws)

        wb.close()  # close the workbook


# creating a array with all the seasons #
year = 1980
seasons = []
while year < 2023:
    seasons.append(Season(str(year)))
    year = year + 1

# creating all teams average arrays
FG_avg = []
FG_attempted_avg = []
FG_rate_avg = []
ThreeP_avg = []
ThreeP_attempted_avg = []
ThreeP_rate_avg = []
TwoP_avg = []
TwoP_attempted_avg = []
TwoP_rate_avg = []
FT_avg = []
FT_attempted_avg = []
FT_rate_avg = []
OR_avg = []
DR_avg = []
AST_avg = []
STL_avg = []
BLK_avg = []
TOV_avg = []
PF_avg = []
PTS_avg = []

for i in seasons:
    FG_avg.append(sum(i.FG) / len(i.FG))
    FG_attempted_avg.append(sum(i.FG_attempted) / len(i.FG_attempted))
    FG_rate_avg.append(sum(i.FG_rate) / len(i.FG_rate))
    ThreeP_avg.append(sum(i.ThreeP) / len(i.ThreeP))
    ThreeP_attempted_avg.append(sum(i.ThreeP_attempted) / len(i.ThreeP_attempted))
    ThreeP_rate_avg.append(sum(i.ThreeP_rate) / len(i.ThreeP_rate))
    TwoP_avg.append(sum(i.TwoP) / len(i.TwoP))
    TwoP_attempted_avg.append(sum(i.TwoP_attempted) / len(i.TwoP_attempted))
    TwoP_rate_avg.append(sum(i.TwoP_rate) / len(i.TwoP_rate))
    FT_avg.append(sum(i.FT) / len(i.FT))
    FT_attempted_avg.append(sum(i.FT_attempted) / len(i.FT_attempted))
    FT_rate_avg.append(sum(i.FT_rate) / len(i.FT_rate))
    OR_avg.append(sum(i.OR) / len(i.OR))
    DR_avg.append(sum(i.DR) / len(i.DR))
    AST_avg.append(sum(i.AST) / len(i.AST))
    STL_avg.append(sum(i.STL) / len(i.STL))
    BLK_avg.append(sum(i.BLK) / len(i.BLK))
    TOV_avg.append(sum(i.TOV) / len(i.TOV))
    PF_avg.append(sum(i.PF) / len(i.PF))
    PTS_avg.append(sum(i.PTS) / len(i.PTS))

# making the avg plots

# making the array with personalized variables
title = ["field goals made", "field goals attempted", "field goal percentage", "three-pointers made", "three-pointers attempted", "three-point percentage", "two-pointers made", "two-pointers attempted", "two-point percentage", "free throws made", "free throws attempted", "free throw percentage", "offensive rebounds", "defensive rebounds", "assists", "steals", "blocks", "turnovers", "personal fouls", "points"]
data = [FG_avg, FG_attempted_avg, FG_rate_avg, ThreeP_avg, ThreeP_attempted_avg, ThreeP_rate_avg, TwoP_avg, TwoP_attempted_avg, TwoP_rate_avg, FT_avg, FT_attempted_avg, FT_rate_avg, OR_avg, DR_avg, AST_avg, STL_avg, BLK_avg, TOV_avg, PF_avg, PTS_avg]

# plot
for a in range(0, len(data)):
    fig, ax = plt.subplots()
    ax.set_title("NBA average " + title[a] + " per game")
    ax.set_xlabel("end season year")
    ax.set_ylabel(title[a] + "per game average")
    ax.plot(range(1980, 2023), data[a], "r+")
    ax.plot(range(1980, 2023), data[a], linewidth=0.7)
    plt.grid()
    plt.xticks(np.arange(1980, 2024, 6))

plt.show()
