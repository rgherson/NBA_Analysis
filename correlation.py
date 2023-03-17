import matplotlib.pyplot
import openpyxl as xl
import matplotlib.pyplot as plt
import numpy as np


# returns a ordered list from a colum ignoring the first element (colum name)
def get_index(column, ws):
    col = []
    for i in range(2, len(ws[column])):
        col.append(float(ws.cell(row=i, column=(ord(column) - 64)).value))
    col.sort()
    return col


# the object that will containg all the data in a season
class Season:
    def __init__(self, workbook):
        self.workbook = workbook
        print("loading workbook: " + workbook)
        wb = xl.load_workbook(filename='workbooks\\' + workbook + ".xlsx")  # load workbook from file
        ws = wb.active

        # define all the indexes into ordered vectors
        self.FG = get_index('E', ws)
        self.FG_attempted = get_index('F', ws)
        self.FG_rate = get_index('G', ws)
        self.ThreeP = get_index('H', ws)
        self.ThreeP_attempted = get_index('I', ws)
        self.ThreeP_rate = get_index('J', ws)
        self.TwoP = get_index('K', ws)
        self.TwoP_attempted = get_index('L', ws)
        self.TwoP_rate = get_index('M', ws)
        self.FT = get_index('N', ws)
        self.FT_attempted = get_index('O', ws)
        self.FT_rate = get_index('P', ws)
        self.OR = get_index('Q', ws)
        self.DR = get_index('R', ws)
        self.AST = get_index('T', ws)
        self.STL = get_index('U', ws)
        self.BLK = get_index('V', ws)
        self.TOV = get_index('W', ws)
        self.PF = get_index('X', ws)
        self.PTS = get_index('Y', ws)

        wb.close()  # close the workbook


# creating a array with all the seasons #
year = 1980
seasons = []
while year < 2023:
    seasons.append(Season(str(year)))
    year = year + 1

# creating all teams average arrays
FG_avg = []
FG_attempted_avg = []
FG_rate_avg = []
ThreeP_avg = []
ThreeP_attempted_avg = []
ThreeP_rate_avg = []
TwoP_avg = []
TwoP_attempted_avg = []
TwoP_rate_avg = []
FT_avg = []
FT_attempted_avg = []
FT_rate_avg = []
OR_avg = []
DR_avg = []
AST_avg = []
STL_avg = []
BLK_avg = []
TOV_avg = []
PF_avg = []
PTS_avg = []

for i in seasons:
    FG_avg.append(sum(i.FG) / len(i.FG))
    FG_attempted_avg.append(sum(i.FG_attempted) / len(i.FG_attempted))
    FG_rate_avg.append(sum(i.FG_rate) / len(i.FG_rate))
    ThreeP_avg.append(sum(i.ThreeP) / len(i.ThreeP))
    ThreeP_attempted_avg.append(sum(i.ThreeP_attempted) / len(i.ThreeP_attempted))
    ThreeP_rate_avg.append(sum(i.ThreeP_rate) / len(i.ThreeP_rate))
    TwoP_avg.append(sum(i.TwoP) / len(i.TwoP))
    TwoP_attempted_avg.append(sum(i.TwoP_attempted) / len(i.TwoP_attempted))
    TwoP_rate_avg.append(sum(i.TwoP_rate) / len(i.TwoP_rate))
    FT_avg.append(sum(i.FT) / len(i.FT))
    FT_attempted_avg.append(sum(i.FT_attempted) / len(i.FT_attempted))
    FT_rate_avg.append(sum(i.FT_rate) / len(i.FT_rate))
    OR_avg.append(sum(i.OR) / len(i.OR))
    DR_avg.append(sum(i.DR) / len(i.DR))
    AST_avg.append(sum(i.AST) / len(i.AST))
    STL_avg.append(sum(i.STL) / len(i.STL))
    BLK_avg.append(sum(i.BLK) / len(i.BLK))
    TOV_avg.append(sum(i.TOV) / len(i.TOV))
    PF_avg.append(sum(i.PF) / len(i.PF))
    PTS_avg.append(sum(i.PTS) / len(i.PTS))


# calculating the correlations
data = [FG_avg, FG_attempted_avg, FG_rate_avg, ThreeP_avg, ThreeP_attempted_avg, ThreeP_rate_avg, TwoP_avg, TwoP_attempted_avg, TwoP_rate_avg, FT_avg, FT_attempted_avg, FT_rate_avg, OR_avg, DR_avg, AST_avg, STL_avg, BLK_avg, TOV_avg, PF_avg, PTS_avg]
name = ["FG_avg", "FG_attempted_avg", "FG_rate_avg", "ThreeP_avg", "ThreeP_attempted_avg", "ThreeP_rate_avg", "TwoP_avg", "TwoP_attempted_avg", "TwoP_rate_avg", "FT_avg", "FT_attempted_avg", "FT_rate_avg", "OR_avg", "DR_avg", "AST_avg", "STL_avg", "BLK_avg", "TOV_avg", "PF_avg", "PTS_avg"]

correlation = np.corrcoef(data)

for i in range(0, 20):
    for k in range(i, 20):
        if correlation[i][k] < 0.99991:
            print("Correlation of " + str(correlation[i][k]) + " between " + name[i] + " and " + name[k])
'''

# ploting the 2 hypothesis

# graph 1
fig1, axs = plt.subplots(2, 2)
fig1.suptitle("NBA hypothesis of correlation between close contact attempts and offencive rebound")

# Three_attempted and OR
axs[0,0].set_title("Correlation = -0.94")
axs[0,0].plot(range(1980, 2023), ThreeP_attempted_avg, "+", color="#16FF00")
axs[0,0].plot(range(1980, 2023), OR_avg, "+", color="#FFED00")
axs[0,0].plot(range(1980, 2023), ThreeP_attempted_avg, linewidth=0.7, color="#000000")
axs[0,0].plot(range(1980, 2023), OR_avg, linewidth=0.7, color="#0F6292")
line1, = axs[0,0].plot([], label='Threes attempted', color="#000000")
line2, = axs[0,0].plot([], label='OR', color="#0F6292")
axs[0,0].legend(handles=[line1, line2], loc="upper right")

# OR and FT_atmp
axs[0,1].set_title("Correlation = 0.91")
axs[0,1].plot(range(1980, 2023), FT_attempted_avg, "+", color="#16FF00")
axs[0,1].plot(range(1980, 2023), OR_avg, "+", color="#FFED00")
axs[0,1].plot(range(1980, 2023), FT_attempted_avg, linewidth=0.7, color="#000000")
axs[0,1].plot(range(1980, 2023), OR_avg, linewidth=0.7, color="#0F6292")
line1, = axs[0,1].plot([], label='FT attempted', color="#000000")
line2, = axs[0,1].plot([], label='OR', color="#0F6292")
axs[0,1].legend(handles=[line1, line2], loc="upper right")

# PF and OR
axs[1,0].set_title("Correlation = 0.88")
axs[1,0].plot(range(1980, 2023), PF_avg, "+", color="#16FF00")
axs[1,0].plot(range(1980, 2023), OR_avg, "+", color="#FFED00")
axs[1,0].plot(range(1980, 2023), PF_avg, linewidth=0.7, color="#000000")
axs[1,0].plot(range(1980, 2023), OR_avg, linewidth=0.7, color="#0F6292")
line1, = axs[1,0].plot([], label='PF', color="#000000")
line2, = axs[1,0].plot([], label='OR', color="#0F6292")
axs[1,0].legend(handles=[line1, line2], loc="upper right")

# Twos_atmp and OR
axs[1,1].set_title("Correlation = 0.95")
axs[1,1].plot(range(1980, 2023), TwoP_attempted_avg, "+", color="#16FF00")
axs[1,1].plot(range(1980, 2023), OR_avg, "+", color="#FFED00")
axs[1,1].plot(range(1980, 2023), TwoP_attempted_avg, linewidth=0.7, color="#000000")
axs[1,1].plot(range(1980, 2023), OR_avg, linewidth=0.7, color="#0F6292")
line1, = axs[1,1].plot([], label='Twos attempted', color="#000000")
line2, = axs[1,1].plot([], label='OR', color="#0F6292")
axs[1,1].legend(handles=[line1, line2], loc="upper right")


# graph 2
fig2, axs = plt.subplots(2, 2)
fig2.suptitle("NBA hypothesis of correlation between close contact attempts and turnovers")

# Three_attempted and TOV
axs[0,0].set_title("Correlation = -0.84")
axs[0,0].plot(range(1980, 2023), ThreeP_attempted_avg, "+", color="#16FF00")
axs[0,0].plot(range(1980, 2023),TOV_avg, "+", color="#FFED00")
axs[0,0].plot(range(1980, 2023), ThreeP_attempted_avg, linewidth=0.7, color="#000000")
axs[0,0].plot(range(1980, 2023), TOV_avg, linewidth=0.7, color="#0F6292")
line1, = axs[0,0].plot([], label='Threes attempted', color="#000000")
line2, = axs[0,0].plot([], label='TOV', color="#0F6292")
axs[0,0].legend(handles=[line1, line2], loc="upper right")

# TOV and FT
axs[0,1].set_title("Correlation = 0.85")
axs[0,1].plot(range(1980, 2023), FT_attempted_avg, "+", color="#16FF00")
axs[0,1].plot(range(1980, 2023), TOV_avg, "+", color="#FFED00")
axs[0,1].plot(range(1980, 2023), FT_attempted_avg, linewidth=0.7, color="#000000")
axs[0,1].plot(range(1980, 2023), TOV_avg, linewidth=0.7, color="#0F6292")
line1, = axs[0,1].plot([], label='FT attempted', color="#000000")
line2, = axs[0,1].plot([], label='TOV', color="#0F6292")
axs[0,1].legend(handles=[line1, line2], loc="upper right")

# PF and TOV
axs[1,0].set_title("Correlation = 0.91")
axs[1,0].plot(range(1980, 2023), PF_avg, "+", color="#16FF00")
axs[1,0].plot(range(1980, 2023), TOV_avg, "+", color="#FFED00")
axs[1,0].plot(range(1980, 2023), PF_avg, linewidth=0.7, color="#000000")
axs[1,0].plot(range(1980, 2023), TOV_avg, linewidth=0.7, color="#0F6292")
line1, = axs[1,0].plot([], label='PF', color="#000000")
line2, = axs[1,0].plot([], label='TOV', color="#0F6292")
axs[1,0].legend(handles=[line1, line2], loc="upper right")

# Twos_atmp and TOV
axs[1,1].set_title("Correlation = 0.92")
axs[1,1].plot(range(1980, 2023), TwoP_attempted_avg, "+", color="#16FF00")
axs[1,1].plot(range(1980, 2023), TOV_avg, "+", color="#FFED00")
axs[1,1].plot(range(1980, 2023), TwoP_attempted_avg, linewidth=0.7, color="#000000")
axs[1,1].plot(range(1980, 2023), TOV_avg, linewidth=0.7, color="#0F6292")
line1, = axs[1,1].plot([], label='Twos attempted', color="#000000")
line2, = axs[1,1].plot([], label='TOV', color="#0F6292")
axs[1,1].legend(handles=[line1, line2], loc="upper right")

# Graph 3
fig3, axs = plt.subplots()
fig3.suptitle("NBA correlation of TOV and OR")
axs.set_title("Correlation = 0.89")
axs.plot(range(1980, 2023), OR_avg, "+", color="#16FF00")
axs.plot(range(1980, 2023), TOV_avg, "+", color="#FFED00")
axs.plot(range(1980, 2023), OR_avg, linewidth=0.7, color="#000000")
axs.plot(range(1980, 2023), TOV_avg, linewidth=0.7, color="#0F6292")
line1, = axs.plot([], label='OR', color="#000000")
line2, = axs.plot([], label='TOV', color="#0F6292")
axs.legend(handles=[line1, line2], loc="upper right")

plt.show()

'''