import matplotlib.pyplot
import openpyxl as xl
import matplotlib.pyplot as plt
import numpy as np


# returns a ordered list from a colum ignoring the first element (colum name)
def get_index(column, ws):
    col = []
    for i in range(2, len(ws[column])):
        col.append(float(ws.cell(row=i, column=(ord(column) - 64)).value))
    col.sort()
    return col


# returns the avg of the first 25% of an array
def low_quartile_avg(target):
    avg = 0
    for k in range(0, round((len(target) + 1) * 0.25)):
        avg = target[k] + avg
    avg = avg / (k+1)

    return avg


# return the top 25% average
def top_quartile_avg(target):
    target.reverse()
    avg = 0
    for k in range(0, round((len(target) + 1) * 0.25)):
        avg = target[k] + avg
    avg = avg / (k+1)

    return avg


# the object that will containg all the data in a season
class Season:
    def __init__(self, workbook):
        self.workbook = workbook
        print("loading workbook: " + workbook)
        wb = xl.load_workbook(filename='workbooks\\' + workbook + ".xlsx")  # load workbook from file
        ws = wb.active

        # define all the indexes into ordered vectors
        self.FG = get_index('E', ws)
        self.FG_attempted = get_index('F', ws)
        self.FG_rate = get_index('G', ws)
        self.ThreeP = get_index('H', ws)
        self.ThreeP_attempted = get_index('I', ws)
        self.ThreeP_rate = get_index('J', ws)
        self.TwoP = get_index('K', ws)
        self.TwoP_attempted = get_index('L', ws)
        self.TwoP_rate = get_index('M', ws)
        self.FT = get_index('N', ws)
        self.FT_attempted = get_index('O', ws)
        self.FT_rate = get_index('P', ws)
        self.OR = get_index('Q', ws)
        self.DR = get_index('R', ws)
        self.AST = get_index('T', ws)
        self.STL = get_index('U', ws)
        self.BLK = get_index('V', ws)
        self.TOV = get_index('W', ws)
        self.PF = get_index('X', ws)
        self.PTS = get_index('Y', ws)

        wb.close()  # close the workbook


# creating a array with all the seasons
year = 1980
seasons = []
while year < 2023:
    seasons.append(Season(str(year)))
    year = year + 1

# creating all teams average arrays
FG_avg = []
FG_attempted_avg = []
FG_rate_avg = []
ThreeP_avg = []
ThreeP_attempted_avg = []
ThreeP_rate_avg = []
TwoP_avg = []
TwoP_attempted_avg = []
TwoP_rate_avg = []
FT_avg = []
FT_attempted_avg = []
FT_rate_avg = []
OR_avg = []
DR_avg = []
AST_avg = []
STL_avg = []
BLK_avg = []
TOV_avg = []
PF_avg = []
PTS_avg = []

for i in seasons:
    FG_avg.append(sum(i.FG) / len(i.FG))
    FG_attempted_avg.append(sum(i.FG_attempted) / len(i.FG_attempted))
    FG_rate_avg.append(sum(i.FG_rate) / len(i.FG_rate))
    ThreeP_avg.append(sum(i.ThreeP) / len(i.ThreeP))
    ThreeP_attempted_avg.append(sum(i.ThreeP_attempted) / len(i.ThreeP_attempted))
    ThreeP_rate_avg.append(sum(i.ThreeP_rate) / len(i.ThreeP_rate))
    TwoP_avg.append(sum(i.TwoP) / len(i.TwoP))
    TwoP_attempted_avg.append(sum(i.TwoP_attempted) / len(i.TwoP_attempted))
    TwoP_rate_avg.append(sum(i.TwoP_rate) / len(i.TwoP_rate))
    FT_avg.append(sum(i.FT) / len(i.FT))
    FT_attempted_avg.append(sum(i.FT_attempted) / len(i.FT_attempted))
    FT_rate_avg.append(sum(i.FT_rate) / len(i.FT_rate))
    OR_avg.append(sum(i.OR) / len(i.OR))
    DR_avg.append(sum(i.DR) / len(i.DR))
    AST_avg.append(sum(i.AST) / len(i.AST))
    STL_avg.append(sum(i.STL) / len(i.STL))
    BLK_avg.append(sum(i.BLK) / len(i.BLK))
    TOV_avg.append(sum(i.TOV) / len(i.TOV))
    PF_avg.append(sum(i.PF) / len(i.PF))
    PTS_avg.append(sum(i.PTS) / len(i.PTS))

# creating low teams average arrays
FG_low_avg = []
FG_attempted_low_avg = []
FG_rate_low_avg = []
ThreeP_low_avg = []
ThreeP_attempted_low_avg = []
ThreeP_rate_low_avg = []
TwoP_low_avg = []
TwoP_attempted_low_avg = []
TwoP_rate_low_avg = []
FT_low_avg = []
FT_attempted_low_avg = []
FT_rate_low_avg = []
OR_low_avg = []
DR_low_avg = []
AST_low_avg = []
STL_low_avg = []
BLK_low_avg = []
TOV_low_avg = []
PF_low_avg = []
PTS_low_avg = []

for i in seasons:
    FG_low_avg.append(low_quartile_avg(i.FG))
    FG_attempted_low_avg.append(low_quartile_avg(i.FG_attempted))
    FG_rate_low_avg.append(low_quartile_avg(i.FG_rate))
    ThreeP_low_avg.append(low_quartile_avg(i.ThreeP))
    ThreeP_attempted_low_avg.append(low_quartile_avg(i.ThreeP_attempted))
    ThreeP_rate_low_avg.append(low_quartile_avg(i.ThreeP_rate))
    TwoP_low_avg.append(low_quartile_avg(i.TwoP))
    TwoP_attempted_low_avg.append(low_quartile_avg(i.TwoP_attempted))
    TwoP_rate_low_avg.append(low_quartile_avg(i.TwoP_rate))
    FT_low_avg.append(low_quartile_avg(i.FT))
    FT_attempted_low_avg.append(low_quartile_avg(i.FT_attempted))
    FT_rate_low_avg.append(low_quartile_avg(i.FT_rate))
    OR_low_avg.append(low_quartile_avg(i.OR))
    DR_low_avg.append(low_quartile_avg(i.DR))
    AST_low_avg.append(low_quartile_avg(i.AST))
    STL_low_avg.append(low_quartile_avg(i.STL))
    BLK_low_avg.append(low_quartile_avg(i.BLK))
    TOV_low_avg.append(low_quartile_avg(i.TOV))
    PF_low_avg.append(low_quartile_avg(i.PF))
    PTS_low_avg.append(low_quartile_avg(i.PTS))

# creating the top 25% avg arrays
# creating top teams average arrays
FG_top_avg = []
FG_attempted_top_avg = []
FG_rate_top_avg = []
ThreeP_top_avg = []
ThreeP_attempted_top_avg = []
ThreeP_rate_top_avg = []
TwoP_top_avg = []
TwoP_attempted_top_avg = []
TwoP_rate_top_avg = []
FT_top_avg = []
FT_attempted_top_avg = []
FT_rate_top_avg = []
OR_top_avg = []
DR_top_avg = []
AST_top_avg = []
STL_top_avg = []
BLK_top_avg = []
TOV_top_avg = []
PF_top_avg = []
PTS_top_avg = []

for i in seasons:
    FG_top_avg.append(top_quartile_avg(i.FG))
    FG_attempted_top_avg.append(top_quartile_avg(i.FG_attempted))
    FG_rate_top_avg.append(top_quartile_avg(i.FG_rate))
    ThreeP_top_avg.append(top_quartile_avg(i.ThreeP))
    ThreeP_attempted_top_avg.append(top_quartile_avg(i.ThreeP_attempted))
    ThreeP_rate_top_avg.append(top_quartile_avg(i.ThreeP_rate))
    TwoP_top_avg.append(top_quartile_avg(i.TwoP))
    TwoP_attempted_top_avg.append(top_quartile_avg(i.TwoP_attempted))
    TwoP_rate_top_avg.append(top_quartile_avg(i.TwoP_rate))
    FT_top_avg.append(top_quartile_avg(i.FT))
    FT_attempted_top_avg.append(top_quartile_avg(i.FT_attempted))
    FT_rate_top_avg.append(top_quartile_avg(i.FT_rate))
    OR_top_avg.append(top_quartile_avg(i.OR))
    DR_top_avg.append(top_quartile_avg(i.DR))
    AST_top_avg.append(top_quartile_avg(i.AST))
    STL_top_avg.append(top_quartile_avg(i.STL))
    BLK_top_avg.append(top_quartile_avg(i.BLK))
    TOV_top_avg.append(top_quartile_avg(i.TOV))
    PF_top_avg.append(top_quartile_avg(i.PF))
    PTS_top_avg.append(top_quartile_avg(i.PTS))


# making the plots

# FG_avg
fig1, ax = plt.subplots()
ax.set_title("NBA average of field goals converted per game")
ax.set_xlabel("end season year")
ax.set_ylabel("field goals converted per game")
ax.plot(range(1980, 2023), FG_avg, "r+")
ax.plot(range(1980, 2023), FG_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), FG_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), FG_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), FG_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), FG_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# FG_attempted
fig2, ax = plt.subplots()
ax.set_title("NBA average of field goals attempted per game")
ax.set_xlabel("end season year")
ax.set_ylabel("field goals attempted per game")
ax.plot(range(1980, 2023), FG_attempted_avg, "r+")
ax.plot(range(1980, 2023), FG_attempted_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), FG_attempted_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), FG_attempted_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), FG_attempted_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), FG_attempted_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# FG_rate
fig3, ax = plt.subplots()
ax.set_title("NBA average of field goals percentage per game")
ax.set_xlabel("end season year")
ax.set_ylabel("field goals percentage per game")
ax.plot(range(1980, 2023), FG_rate_avg, "r+")
ax.plot(range(1980, 2023), FG_rate_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), FG_rate_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), FG_rate_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), FG_rate_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), FG_rate_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# ThreeP
fig4, ax = plt.subplots()
ax.set_title("NBA average of three-pointers converted per game")
ax.set_xlabel("end season year")
ax.set_ylabel("three-pointers converted per game")
ax.plot(range(1980, 2023), ThreeP_avg, "r+")
ax.plot(range(1980, 2023), ThreeP_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), ThreeP_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), ThreeP_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), ThreeP_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), ThreeP_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# ThreeP_attempted_avg
fig5, ax = plt.subplots()
ax.set_title("NBA average of three-pointers attempted per game")
ax.set_xlabel("end season year")
ax.set_ylabel("three-pointers attempted per game")
ax.plot(range(1980, 2023), ThreeP_attempted_avg, "r+")
ax.plot(range(1980, 2023), ThreeP_attempted_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), ThreeP_attempted_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), ThreeP_attempted_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), ThreeP_attempted_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), ThreeP_attempted_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# ThreeP_rate_avg
fig6, ax = plt.subplots()
ax.set_title("NBA average of three-pointers percentage per game")
ax.set_xlabel("end season year")
ax.set_ylabel("three-pointers percentage per game")
ax.plot(range(1980, 2023), ThreeP_rate_avg, "r+")
ax.plot(range(1980, 2023), ThreeP_rate_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), ThreeP_rate_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), ThreeP_rate_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), ThreeP_rate_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), ThreeP_rate_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# TwoP_avg
fig7, ax = plt.subplots()
ax.set_title("NBA average of two-pointers converted per game")
ax.set_xlabel("end season year")
ax.set_ylabel("two-pointers converted per game")
ax.plot(range(1980, 2023), TwoP_avg, "r+")
ax.plot(range(1980, 2023), TwoP_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), TwoP_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), TwoP_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), TwoP_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), TwoP_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# TwoP_attempted_avg
fig8, ax = plt.subplots()
ax.set_title("NBA average of two-pointers attempted per game")
ax.set_xlabel("end season year")
ax.set_ylabel("two-pointers attempted per game")
ax.plot(range(1980, 2023), TwoP_attempted_avg, "r+")
ax.plot(range(1980, 2023), TwoP_attempted_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), TwoP_attempted_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), TwoP_attempted_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), TwoP_attempted_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), TwoP_attempted_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# TwoP_rate_avg
fig9, ax = plt.subplots()
ax.set_title("NBA average of two-pointers percentage per game")
ax.set_xlabel("end season year")
ax.set_ylabel("two-pointers percentage per game")
ax.plot(range(1980, 2023), TwoP_rate_avg, "r+")
ax.plot(range(1980, 2023), TwoP_rate_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), TwoP_rate_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), TwoP_rate_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), TwoP_rate_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), TwoP_rate_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# FT_avg
fig10, ax = plt.subplots()
ax.set_title("NBA average of free throws made per game")
ax.set_xlabel("end season year")
ax.set_ylabel("free throws per game")
ax.plot(range(1980, 2023), FT_avg, "r+")
ax.plot(range(1980, 2023), FT_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), FT_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), FT_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), FT_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), FT_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# FT_attempted_avg
fig11, ax = plt.subplots()
ax.set_title("NBA average of free throws attempted per game")
ax.set_xlabel("end season year")
ax.set_ylabel("free throws attempted per game average")
ax.plot(range(1980, 2023), FT_attempted_avg, "r+")
ax.plot(range(1980, 2023), FT_attempted_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), FT_attempted_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), FT_attempted_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), FT_attempted_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), FT_attempted_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# FT_rate_avg
fig12, ax = plt.subplots()
ax.set_title("NBA average of free throws percentage per game")
ax.set_xlabel("end season year")
ax.set_ylabel("free throws percentage per game average")
ax.plot(range(1980, 2023), FT_rate_avg, "r+")
ax.plot(range(1980, 2023), FT_rate_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), FT_rate_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), FT_rate_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), FT_rate_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), FT_rate_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# offensive rebounds per game
fig13, ax = plt.subplots()
ax.set_title("NBA average of offensive rebounds per game per game")
ax.set_xlabel("end season year")
ax.set_ylabel("offensive rebounds per game average")
ax.plot(range(1980, 2023), OR_avg, "r+")
ax.plot(range(1980, 2023), OR_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), OR_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), OR_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), OR_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), OR_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# DR_avg
fig14, ax = plt.subplots()
ax.set_title("NBA average of defensive rebounds per game per game")
ax.set_xlabel("end season year")
ax.set_ylabel("defensive rebounds per game average")
ax.plot(range(1980, 2023), DR_avg, "r+")
ax.plot(range(1980, 2023), DR_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), DR_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), DR_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), DR_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), DR_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# AST_avg
fig15, ax = plt.subplots()
ax.set_title("NBA average of assists per game per game")
ax.set_xlabel("end season year")
ax.set_ylabel("assists per game average")
ax.plot(range(1980, 2023), AST_avg, "r+")
ax.plot(range(1980, 2023), AST_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), AST_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), AST_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), AST_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), AST_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# STL_avg
fig16, ax = plt.subplots()
ax.set_title("NBA average of steals per game")
ax.set_xlabel("end season year")
ax.set_ylabel("steals per game average")
ax.plot(range(1980, 2023), STL_avg, "r+")
ax.plot(range(1980, 2023), STL_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), STL_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), STL_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), STL_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), STL_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# BLK_avg
fig17, ax = plt.subplots()
ax.set_title("NBA average of blocks per game")
ax.set_xlabel("end season year")
ax.set_ylabel("blocks per game average")
ax.plot(range(1980, 2023), BLK_avg, "r+")
ax.plot(range(1980, 2023), BLK_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), BLK_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), BLK_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), BLK_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), BLK_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# TOV_avg
fig18, ax = plt.subplots()
ax.set_title("NBA average of turnovers per game")
ax.set_xlabel("end season year")
ax.set_ylabel("turnovers per game")
ax.plot(range(1980, 2023), TOV_avg, "r+")
ax.plot(range(1980, 2023), TOV_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), TOV_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), TOV_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), TOV_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), TOV_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# PF_avg
fig19, ax = plt.subplots()
ax.set_title("NBA average of personal fouls per game")
ax.set_xlabel("end season year")
ax.set_ylabel("personal fouls per game")
ax.plot(range(1980, 2023), PF_avg, "r+")
ax.plot(range(1980, 2023), PF_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), PF_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), PF_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), PF_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), PF_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

# PTS_avg
fig20, ax = plt.subplots()
ax.set_title("NBA average of points per game")
ax.set_xlabel("end season year")
ax.set_ylabel("points per game")
ax.plot(range(1980, 2023), PTS_avg, "r+")
ax.plot(range(1980, 2023), PTS_avg, linewidth=0.7, color="#5ca0d3")
ax.plot(range(1980, 2023), PTS_low_avg, "+", color="#F39189")
ax.plot(range(1980, 2023), PTS_low_avg, "g", linewidth=0.7)
ax.plot(range(1980, 2023), PTS_top_avg, "+", linewidth=0.7, color="#47555e")
ax.plot(range(1980, 2023), PTS_top_avg, color="#5e366a", linewidth=0.7)
line1, = ax.plot([], label='all teams', color="#5ca0d3")
line2, = ax.plot([], label='lower 25%', color="g")
line3, = ax.plot([], label='upper 25%', color="#5e366a")
ax.legend(handles=[line1, line2, line3], loc="upper right")
plt.grid()
plt.xticks(np.arange(1980, 2024, 6))

plt.show()
