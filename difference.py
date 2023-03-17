import matplotlib.pyplot
import openpyxl as xl
import matplotlib.pyplot as plt
import numpy as np


# returns a ordered list from a colum ignoring the first element (colum name)
def get_index(column, ws):
    col = []
    for i in range(2, len(ws[column])):
        col.append(float(ws.cell(row=i, column=(ord(column) - 64)).value))
    col.sort()
    return col


# returns the avg of the first 25% of an array
def low_quartile_avg(target):
    avg = 0
    for k in range(0, round((len(target) + 1) * 0.25)):
        avg = target[k] + avg
    avg = avg / (k+1)

    return avg


# return the top 25% average
def top_quartile_avg(target):
    target.reverse()
    avg = 0
    for k in range(0, round((len(target) + 1) * 0.25)):
        avg = target[k] + avg
    avg = avg / (k+1)

    return avg


# the object that will containg all the data in a season
class Season:
    def __init__(self, workbook):
        self.workbook = workbook
        print("loading workbook: " + workbook)
        wb = xl.load_workbook(filename='workbooks\\' + workbook + ".xlsx")  # load workbook from file
        ws = wb.active

        # define all the indexes into ordered vectors
        self.FG = get_index('E', ws)
        self.FG_attempted = get_index('F', ws)
        self.FG_rate = get_index('G', ws)
        self.ThreeP = get_index('H', ws)
        self.ThreeP_attempted = get_index('I', ws)
        self.ThreeP_rate = get_index('J', ws)
        self.TwoP = get_index('K', ws)
        self.TwoP_attempted = get_index('L', ws)
        self.TwoP_rate = get_index('M', ws)
        self.FT = get_index('N', ws)
        self.FT_attempted = get_index('O', ws)
        self.FT_rate = get_index('P', ws)
        self.OR = get_index('Q', ws)
        self.DR = get_index('R', ws)
        self.AST = get_index('T', ws)
        self.STL = get_index('U', ws)
        self.BLK = get_index('V', ws)
        self.TOV = get_index('W', ws)
        self.PF = get_index('X', ws)
        self.PTS = get_index('Y', ws)

        wb.close()  # close the workbook


# creating a array with all the seasons
year = 1980
seasons = []
while year < 2023:
    seasons.append(Season(str(year)))
    year = year + 1

# creating all teams average arrays
FG_avg = []
FG_attempted_avg = []
FG_rate_avg = []
ThreeP_avg = []
ThreeP_attempted_avg = []
ThreeP_rate_avg = []
TwoP_avg = []
TwoP_attempted_avg = []
TwoP_rate_avg = []
FT_avg = []
FT_attempted_avg = []
FT_rate_avg = []
OR_avg = []
DR_avg = []
AST_avg = []
STL_avg = []
BLK_avg = []
TOV_avg = []
PF_avg = []
PTS_avg = []

for i in seasons:
    FG_avg.append(sum(i.FG) / len(i.FG))
    FG_attempted_avg.append(sum(i.FG_attempted) / len(i.FG_attempted))
    FG_rate_avg.append(sum(i.FG_rate) / len(i.FG_rate))
    ThreeP_avg.append(sum(i.ThreeP) / len(i.ThreeP))
    ThreeP_attempted_avg.append(sum(i.ThreeP_attempted) / len(i.ThreeP_attempted))
    ThreeP_rate_avg.append(sum(i.ThreeP_rate) / len(i.ThreeP_rate))
    TwoP_avg.append(sum(i.TwoP) / len(i.TwoP))
    TwoP_attempted_avg.append(sum(i.TwoP_attempted) / len(i.TwoP_attempted))
    TwoP_rate_avg.append(sum(i.TwoP_rate) / len(i.TwoP_rate))
    FT_avg.append(sum(i.FT) / len(i.FT))
    FT_attempted_avg.append(sum(i.FT_attempted) / len(i.FT_attempted))
    FT_rate_avg.append(sum(i.FT_rate) / len(i.FT_rate))
    OR_avg.append(sum(i.OR) / len(i.OR))
    DR_avg.append(sum(i.DR) / len(i.DR))
    AST_avg.append(sum(i.AST) / len(i.AST))
    STL_avg.append(sum(i.STL) / len(i.STL))
    BLK_avg.append(sum(i.BLK) / len(i.BLK))
    TOV_avg.append(sum(i.TOV) / len(i.TOV))
    PF_avg.append(sum(i.PF) / len(i.PF))
    PTS_avg.append(sum(i.PTS) / len(i.PTS))

# creating low teams average arrays
FG_low_avg = []
FG_attempted_low_avg = []
FG_rate_low_avg = []
ThreeP_low_avg = []
ThreeP_attempted_low_avg = []
ThreeP_rate_low_avg = []
TwoP_low_avg = []
TwoP_attempted_low_avg = []
TwoP_rate_low_avg = []
FT_low_avg = []
FT_attempted_low_avg = []
FT_rate_low_avg = []
OR_low_avg = []
DR_low_avg = []
AST_low_avg = []
STL_low_avg = []
BLK_low_avg = []
TOV_low_avg = []
PF_low_avg = []
PTS_low_avg = []

for i in seasons:
    FG_low_avg.append(low_quartile_avg(i.FG))
    FG_attempted_low_avg.append(low_quartile_avg(i.FG_attempted))
    FG_rate_low_avg.append(low_quartile_avg(i.FG_rate))
    ThreeP_low_avg.append(low_quartile_avg(i.ThreeP))
    ThreeP_attempted_low_avg.append(low_quartile_avg(i.ThreeP_attempted))
    ThreeP_rate_low_avg.append(low_quartile_avg(i.ThreeP_rate))
    TwoP_low_avg.append(low_quartile_avg(i.TwoP))
    TwoP_attempted_low_avg.append(low_quartile_avg(i.TwoP_attempted))
    TwoP_rate_low_avg.append(low_quartile_avg(i.TwoP_rate))
    FT_low_avg.append(low_quartile_avg(i.FT))
    FT_attempted_low_avg.append(low_quartile_avg(i.FT_attempted))
    FT_rate_low_avg.append(low_quartile_avg(i.FT_rate))
    OR_low_avg.append(low_quartile_avg(i.OR))
    DR_low_avg.append(low_quartile_avg(i.DR))
    AST_low_avg.append(low_quartile_avg(i.AST))
    STL_low_avg.append(low_quartile_avg(i.STL))
    BLK_low_avg.append(low_quartile_avg(i.BLK))
    TOV_low_avg.append(low_quartile_avg(i.TOV))
    PF_low_avg.append(low_quartile_avg(i.PF))
    PTS_low_avg.append(low_quartile_avg(i.PTS))

# creating the top 25% avg arrays
# creating top teams average arrays
FG_top_avg = []
FG_attempted_top_avg = []
FG_rate_top_avg = []
ThreeP_top_avg = []
ThreeP_attempted_top_avg = []
ThreeP_rate_top_avg = []
TwoP_top_avg = []
TwoP_attempted_top_avg = []
TwoP_rate_top_avg = []
FT_top_avg = []
FT_attempted_top_avg = []
FT_rate_top_avg = []
OR_top_avg = []
DR_top_avg = []
AST_top_avg = []
STL_top_avg = []
BLK_top_avg = []
TOV_top_avg = []
PF_top_avg = []
PTS_top_avg = []

for i in seasons:
    FG_top_avg.append(top_quartile_avg(i.FG))
    FG_attempted_top_avg.append(top_quartile_avg(i.FG_attempted))
    FG_rate_top_avg.append(top_quartile_avg(i.FG_rate))
    ThreeP_top_avg.append(top_quartile_avg(i.ThreeP))
    ThreeP_attempted_top_avg.append(top_quartile_avg(i.ThreeP_attempted))
    ThreeP_rate_top_avg.append(top_quartile_avg(i.ThreeP_rate))
    TwoP_top_avg.append(top_quartile_avg(i.TwoP))
    TwoP_attempted_top_avg.append(top_quartile_avg(i.TwoP_attempted))
    TwoP_rate_top_avg.append(top_quartile_avg(i.TwoP_rate))
    FT_top_avg.append(top_quartile_avg(i.FT))
    FT_attempted_top_avg.append(top_quartile_avg(i.FT_attempted))
    FT_rate_top_avg.append(top_quartile_avg(i.FT_rate))
    OR_top_avg.append(top_quartile_avg(i.OR))
    DR_top_avg.append(top_quartile_avg(i.DR))
    AST_top_avg.append(top_quartile_avg(i.AST))
    STL_top_avg.append(top_quartile_avg(i.STL))
    BLK_top_avg.append(top_quartile_avg(i.BLK))
    TOV_top_avg.append(top_quartile_avg(i.TOV))
    PF_top_avg.append(top_quartile_avg(i.PF))
    PTS_top_avg.append(top_quartile_avg(i.PTS))

# creat now the array of the diference betwen two seasons
FG_top_dif = []
FG_attempted_top_dif = []
FG_rate_top_dif = []
ThreeP_top_dif = []
ThreeP_attempted_top_dif = []
ThreeP_rate_top_dif = []
TwoP_top_dif = []
TwoP_attempted_top_dif = []
TwoP_rate_top_dif = []
FT_top_dif = []
FT_attempted_top_dif = []
FT_rate_top_dif = []
OR_top_dif = []
DR_top_dif = []
AST_top_dif = []
STL_top_dif = []
BLK_top_dif = []
TOV_top_dif = []
PF_top_dif = []
PTS_top_dif = []
FG_low_dif = []
FG_attempted_low_dif = []
FG_rate_low_dif = []
ThreeP_low_dif = []
ThreeP_attempted_low_dif = []
ThreeP_rate_low_dif = []
TwoP_low_dif = []
TwoP_attempted_low_dif = []
TwoP_rate_low_dif = []
FT_low_dif = []
FT_attempted_low_dif = []
FT_rate_low_dif = []
OR_low_dif = []
DR_low_dif = []
AST_low_dif = []
STL_low_dif = []
BLK_low_dif = []
TOV_low_dif = []
PF_low_dif = []
PTS_low_dif = []
FG_dif = []
FG_attempted_dif = []
FG_rate_dif = []
ThreeP_dif = []
ThreeP_attempted_dif = []
ThreeP_rate_dif = []
TwoP_dif = []
TwoP_attempted_dif = []
TwoP_rate_dif = []
FT_dif = []
FT_attempted_dif = []
FT_rate_dif = []
OR_dif = []
DR_dif = []
AST_dif = []
STL_dif = []
BLK_dif = []
TOV_dif = []
PF_dif = []
PTS_dif = []

for i in range(0, len(seasons) - 1):
    FG_top_dif.append(FG_top_avg[i+1] - FG_top_avg[i])
    FG_attempted_top_dif.append(FG_attempted_top_avg[i+1] - FG_attempted_top_avg[i])
    FG_rate_top_dif.append(FG_rate_top_avg[i+1] - FG_rate_top_avg[i])
    ThreeP_top_dif.append(ThreeP_top_avg[i + 1] - ThreeP_top_avg[i])
    ThreeP_attempted_top_dif.append(ThreeP_attempted_top_avg[i + 1] - ThreeP_attempted_top_avg[i])
    ThreeP_rate_top_dif.append(ThreeP_rate_top_avg[i + 1] - ThreeP_rate_top_avg[i])
    TwoP_top_dif.append(TwoP_top_avg[i + 1] - TwoP_top_avg[i])
    TwoP_attempted_top_dif.append(TwoP_attempted_top_avg[i + 1] - TwoP_attempted_top_avg[i])
    TwoP_rate_top_dif.append(TwoP_rate_top_avg[i + 1] - TwoP_rate_top_avg[i])
    FT_top_dif.append(FT_top_avg[i + 1] - FT_top_avg[i])
    FT_attempted_top_dif.append(FT_attempted_top_avg[i + 1] - FT_attempted_top_avg[i])
    FT_rate_top_dif.append(FT_rate_top_avg[i + 1] - FT_rate_top_avg[i])
    OR_top_dif.append(OR_top_avg[i + 1] - OR_top_avg[i])
    DR_top_dif.append(DR_top_avg[i + 1] - DR_top_avg[i])
    AST_top_dif.append(AST_top_avg[i + 1] - AST_top_avg[i])
    STL_top_dif.append(STL_top_avg[i + 1] - STL_top_avg[i])
    BLK_top_dif.append(BLK_top_avg[i + 1] - BLK_top_avg[i])
    TOV_top_dif.append(TOV_top_avg[i + 1] - TOV_top_avg[i])
    PF_top_dif.append(PF_top_avg[i + 1] - PF_top_avg[i])
    PTS_top_dif.append(PTS_top_avg[i + 1] - PTS_top_avg[i])
    FG_low_dif.append(FG_low_avg[i + 1] - FG_low_avg[i])
    FG_attempted_low_dif.append(FG_attempted_low_avg[i + 1] - FG_attempted_low_avg[i])
    FG_rate_low_dif.append(FG_rate_low_avg[i + 1] - FG_rate_low_avg[i])
    ThreeP_low_dif.append(ThreeP_low_avg[i + 1] - ThreeP_low_avg[i])
    ThreeP_attempted_low_dif.append(ThreeP_attempted_low_avg[i + 1] - ThreeP_attempted_low_avg[i])
    ThreeP_rate_low_dif.append(ThreeP_rate_low_avg[i + 1] - ThreeP_rate_low_avg[i])
    TwoP_low_dif.append(TwoP_low_avg[i + 1] - TwoP_low_avg[i])
    TwoP_attempted_low_dif.append(TwoP_attempted_low_avg[i + 1] - TwoP_attempted_low_avg[i])
    TwoP_rate_low_dif.append(TwoP_rate_low_avg[i + 1] - TwoP_rate_low_avg[i])
    FT_low_dif.append(FT_low_avg[i + 1] - FT_low_avg[i])
    FT_attempted_low_dif.append(FT_attempted_low_avg[i + 1] - FT_attempted_low_avg[i])
    FT_rate_low_dif.append(FT_rate_low_avg[i + 1] - FT_rate_low_avg[i])
    OR_low_dif.append(OR_low_avg[i + 1] - OR_low_avg[i])
    DR_low_dif.append(DR_low_avg[i + 1] - DR_low_avg[i])
    AST_low_dif.append(AST_low_avg[i + 1] - AST_low_avg[i])
    STL_low_dif.append(STL_low_avg[i + 1] - STL_low_avg[i])
    BLK_low_dif.append(BLK_low_avg[i + 1] - BLK_low_avg[i])
    TOV_low_dif.append(TOV_low_avg[i + 1] - TOV_low_avg[i])
    PF_low_dif.append(PF_low_avg[i + 1] - PF_low_avg[i])
    PTS_low_dif.append(PTS_low_avg[i + 1] - PTS_low_avg[i])
    FG_dif.append(FG_avg[i + 1] - FG_avg[i])
    FG_attempted_dif.append(FG_attempted_avg[i + 1] - FG_attempted_avg[i])
    FG_rate_dif.append(FG_rate_avg[i + 1] - FG_rate_avg[i])
    ThreeP_dif.append(ThreeP_avg[i + 1] - ThreeP_avg[i])
    ThreeP_attempted_dif.append(ThreeP_attempted_avg[i + 1] - ThreeP_attempted_avg[i])
    ThreeP_rate_dif.append(ThreeP_rate_avg[i + 1] - ThreeP_rate_avg[i])
    TwoP_dif.append(TwoP_avg[i + 1] - TwoP_avg[i])
    TwoP_attempted_dif.append(TwoP_attempted_avg[i + 1] - TwoP_attempted_avg[i])
    TwoP_rate_dif.append(TwoP_rate_avg[i + 1] - TwoP_rate_avg[i])
    FT_dif.append(FT_avg[i + 1] - FT_avg[i])
    FT_attempted_dif.append(FT_attempted_avg[i + 1] - FT_attempted_avg[i])
    FT_rate_dif.append(FT_rate_avg[i + 1] - FT_rate_avg[i])
    OR_dif.append(OR_avg[i + 1] - OR_avg[i])
    DR_dif.append(DR_avg[i + 1] - DR_avg[i])
    AST_dif.append(AST_avg[i + 1] - AST_avg[i])
    STL_dif.append(STL_avg[i + 1] - STL_avg[i])
    BLK_dif.append(BLK_avg[i + 1] - BLK_avg[i])
    TOV_dif.append(TOV_avg[i + 1] - TOV_avg[i])
    PF_dif.append(PF_avg[i + 1] - PF_avg[i])
    PTS_dif.append(PTS_avg[i + 1] - PTS_avg[i])

# plotting

# organizing everithig that varies in te plot in arrays to be run thought
top = [FG_top_dif, FG_attempted_top_dif, FG_rate_top_dif, ThreeP_top_dif, ThreeP_attempted_top_dif, ThreeP_rate_top_dif, TwoP_top_dif, TwoP_attempted_top_dif, TwoP_rate_top_dif, FT_top_dif, FT_attempted_top_dif, FT_rate_top_dif, OR_top_dif, DR_top_dif, AST_top_dif, STL_top_dif, BLK_top_dif, TOV_top_dif, PF_top_dif]
low = [FG_low_dif, FG_attempted_low_dif, FG_rate_low_dif, ThreeP_low_dif, ThreeP_attempted_low_dif, ThreeP_rate_low_dif, TwoP_low_dif, TwoP_attempted_low_dif, TwoP_rate_low_dif, FT_low_dif, FT_attempted_low_dif, FT_rate_low_dif, OR_low_dif, DR_low_dif, AST_low_dif, STL_low_dif, BLK_low_dif, TOV_low_dif, PF_low_dif]
everyone = [FG_dif, FG_attempted_dif, FG_rate_dif, ThreeP_dif, ThreeP_attempted_dif, ThreeP_rate_dif, TwoP_dif, TwoP_attempted_dif, TwoP_rate_dif, FT_dif, FT_attempted_dif, FT_rate_dif, OR_dif, DR_dif, AST_dif, STL_dif, BLK_dif, TOV_dif, PF_dif]
title = ["field goals converted", "field goals attempted", "field goal percentage", "three-pointers converted", "three-pointers attempted", "three-point percentage", "two-pointers converted", "two-pointers attempted", "two-point percentage", "free throws converted", "free throws attempted", "free throw percentage", "offensive rebounds", "defensive rebounds", "assists", "steals", "blocks", "turnovers", "personal fouls", "points"]

# now actualy ploting it
for i in range(0, len(top)):
    fig, ax = plt.subplots()
    ax.set_title("NBA season to season difference of " + title[i] + " per game")
    ax.set_xlabel("end season year")
    ax.set_ylabel("difference of " + title[i])
    ax.plot(range(1981, 2023), everyone[i], "r+")
    ax.plot(range(1981, 2023), everyone[i], linewidth=0.7, color="#5ca0d3")
    ax.plot(range(1981, 2023), low[i], "+", color="#F39189")
    ax.plot(range(1981, 2023), low[i], "g", linewidth=0.7)
    ax.plot(range(1981, 2023), top[i], "+", linewidth=0.7, color="#47555e")
    ax.plot(range(1981, 2023), top[i], color="#5e366a", linewidth=0.7)
    line1, = ax.plot([], label='all teams', color="#5ca0d3")
    line2, = ax.plot([], label='lower 25%', color="g")
    line3, = ax.plot([], label='upper 25%', color="#5e366a")
    ax.legend(handles=[line1, line2, line3], loc="upper right")
    plt.grid()
    plt.xticks(np.arange(1980, 2024, 6))

plt.show()